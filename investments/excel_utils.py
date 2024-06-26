import sys

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from utils.settings import BASE_DIR


class ModifyExcelFile:

    @staticmethod
    def append_separator(filename: str,
                         cell: str,
                         text: str,
                         font_size: int = 12,
                         fill_color: str = None,
                         fill_pattern: str = None) -> None:
        """
        Записывает текст в указанную ячейку в файле Excel.

        :param filename: Путь к файлу
        :param cell: Ячейка, в которую будет записан текст. Напр. "A1"
        :param text: Текст, который будет записан в ячейку
        :param font_size: Размер шрифта. По умолчанию 12
        :param fill_color: Цвет фона ячейки в формате RGB. По умолчанию None
        :param fill_pattern: Тип заполнения фона ячейки: 'solid' или 'gray125'
        По умолчанию None
        :return: None
        """

        wb = openpyxl.load_workbook(filename)
        ws = wb.active
        ws[cell] = text
        ws[cell].font = Font(size=font_size)
        if fill_color:
            fill = PatternFill(fill_type=fill_pattern, fgColor=fill_color)
            ws[cell].fill = fill
        wb.save(filename)

    @staticmethod
    def auto_fit_columns(filename: str) -> None:
        """
        Выравнивает ширину столбцов по содержимому столбца.

        :param filename: Путь к файлу
        выровнять ширину столбцов
        :return: None
        """

        WIDTH_INCREASE = 1.2
        ADJUST = 2

        wb = openpyxl.load_workbook(filename)

        worksheets = wb.sheetnames

        for sheet_name in worksheets:
            for column in wb[sheet_name].columns:
                max_length = 0
                col_let = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
                adjust_width = (max_length + ADJUST) * WIDTH_INCREASE
                wb[sheet_name].column_dimensions[col_let].width = adjust_width
        wb.save(filename)


class BuildChart:
    def __init__(self, filename: str, sheet_name: str, data_col_num: int):
        self.__filename = filename
        self.__wb = openpyxl.load_workbook(filename)
        self.__ws = self.__wb[sheet_name]
        self.__data_col_num = data_col_num

    def bar_chart(self, chart_labels: list[tuple[str, str, str]]) -> None:
        """
        Построить столюцатую диаграмму. Данные диаграммы и
        количество диапазонов с данными должны совпадать.

        :param chart_labels: Подписи для диаграммы: название диаграммы и
        названия осей x, y. Пример:
        [("Название 1", "Ось х", "Ось y"), ("Название 2", "Ось х", "Ось y")]
        :return: None
        """
        curr_chart_coord = 1
        last_col_letter: str = get_column_letter(self.__ws.max_column + 2)
        data_ranges: list[tuple[int, int]] = self.__get_rows_data_ranges()

        if len(data_ranges) != len(chart_labels):
            sys.stdout.write(
                "Не удалось построить диаграммы! "
                "Удалите полученный файл и попробуйте "
                "снова запустить программу\n"
            )
            return

        for labels, ranges in zip(chart_labels, data_ranges):
            title, x_label, y_label = labels
            start_row, end_row = ranges

            chart = BarChart()
            chart.title = title
            chart.x_axis.title = x_label
            chart.y_axis.title = y_label
            chart.legend = None

            data = Reference(self.__ws,
                             min_col=self.__data_col_num,
                             min_row=start_row,
                             max_col=self.__data_col_num,
                             max_row=end_row)
            chart.add_data(data)

            self.__ws.add_chart(chart, f"{last_col_letter}{curr_chart_coord}")
            curr_chart_coord += 15
            self.__wb.save(self.__filename)

    def get_filename(self):
        return self.__filename

    def __get_rows_data_ranges(self) -> list[tuple[int, int]]:
        """
        Определяет диапазоны ячеек, которые содержат данные.
        :return: list[tuple[int, int]]
        """
        data_ranges: list[tuple[int, int]] = []
        start_row: int | None = None
        max_row: int = self.__ws.max_row

        for row in range(1, max_row + 1):
            cell = self.__ws.cell(row=row,
                                  column=self.__data_col_num).value
            if cell is not None:
                if start_row is None:
                    start_row = row
            elif start_row is not None:
                end_row = row - 1
                data_ranges.append((start_row, end_row))
                start_row = None

        if start_row is not None:
            data_ranges.append((start_row, max_row))

        return data_ranges


if __name__ == '__main__':
    FILE_PATH = rf"{BASE_DIR}\investments\variant_7_data.xlsx"

    ModifyExcelFile.append_separator(
        FILE_PATH,
        "A50",
        "BlaBla"
    )

    ModifyExcelFile.auto_fit_columns(FILE_PATH)
